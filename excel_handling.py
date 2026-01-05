from selenium import webdriver
import openpyxl

# file_=r'D:\PycharmProjects\automation_practise\assets\Book1.xlsx'
# workbook_=openpyxl.load_workbook(file_)
# print([[cell.value for cell in row] for worksheet_ in workbook_ for row in worksheet_])
# for worksheet_ in workbook_:
#     print(list(worksheet_.values))


# worksheet=workbook_.active
# print(worksheet['A1'].value)
# print(worksheet.cell(2,3).value)
# worksheet.max_row

# for r in range(1, worksheet.max_row + 1):
#     for c in range(1, worksheet.max_column + 1):
#         print(worksheet.cell(row=r, column=c).value, end=" | ")
#     print()


import os

if os.path.exists("book.xlsx"):
    print("File exists!")
else:
    print("Not found!")
    
####################################
# writing data into excel
# workbook_=openpyxl.Workbook()
# worksheet_=workbook_.active
# worksheet_.title='info'
'''
# worksheet_['A1']='SlNO'
# worksheet_['B1']='Name'
# worksheet_.append(['SlNo','Name','Age',"Qualification"])'''

# data_=[
#     ['SlNo','Name','Age',"Qualification"],
#     [1,'Lavanya','24','Bsc'],
#     [2,'Bhargav',26,'B.tech']
# ]
# for data in data_:
#     worksheet_.append(data)
# workbook_.save('book1.xlsx')

###############################
# adding new sheet to same workbook
from openpyxl import load_workbook
book=load_workbook('book1.xlsx')
other=book.create_sheet('other info')

data_=[
    ['SlNo','Name','Age',"Qualification"],
    [1,'Lavanya','24','Bsc'],
    [2,'Bhargav',26,'B.tech']
]

for data in data_:
    other.append(data)
book.save('book1.xlsx')

print([[cell.value for cell in row] for other in book for row in other])
